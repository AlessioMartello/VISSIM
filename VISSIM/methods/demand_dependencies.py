import pathlib

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from .helpers import data_inputs_path, project
from .helpers import load_VISSIM_file, df_writer, check_project_name


def get_demand_dependencies(data_directory):
    """ Calculates the number of times the specified demand dependant stages are called. """
    # Read the Excel file, which was manually setup, from the selected data directory, extract the SC number and SCJ
    # number to inform the name of the demand dependant stage.
    dd_file_path = pathlib.Path(data_directory) / 'Demand_dependancy.xlsx'
    book = openpyxl.load_workbook(dd_file_path)
    worksheet = book.active
    sites, sc = [], []
    for row in worksheet.iter_rows(min_row=5):
        site = row[0].value
        signal_control = row[4].value
        if type(site) is int:
            sites.append(site)
            sc.append(signal_control)
        else:
            continue

    # Get times from Excel file to inform where to slice the following DataFrame
    WARMUP_TIME = worksheet.cell(row=5, column=15).value
    COOLDOWN_TIME = worksheet.cell(row=5, column=16).value

    file_number = 0  # initialise variable to count the number of .lsa files
    INTRO_LINES = 8  # Hard code the number of introductory lines to be ignored when reading DataFrame
    SUFFIX = ".lsa"  # The SUFFIX for demand dependency files from VISSIM, to be checked for when reading files
    ASPECT = "green"
    all_results = pd.DataFrame()

    # Read each .lsa file in directory, filter based on the warmup and cooldown times.
    # Get each demand dependant stage count per filtered file.
    # Append to the all_results DataFrame.
    for path in pathlib.Path(data_directory).iterdir():
        if str(path).endswith(SUFFIX):
            raw_data = load_VISSIM_file(path=path, sep=";", columns=["Time", "SCJ", "SC", "Signal"],
                                        use_cols=[0, 2, 3, 4],
                                        skiprows=INTRO_LINES)
            signal_group_list_length = len(raw_data[raw_data["Time"].str.contains("SC")])
            raw_data = load_VISSIM_file(path=path, sep=";", columns=["Time", "SCJ", "SC", "Signal"],
                                        use_cols=[0, 2, 3, 4],
                                        skiprows=INTRO_LINES + signal_group_list_length)
            df = raw_data[(raw_data.Time > WARMUP_TIME) & (raw_data.Time < COOLDOWN_TIME)]
            seed_results = []
            for SCJ_number, SC_number in zip(sites, sc):
                if df.empty:
                    seed_results.append(0)
                else:
                    green_signal_instances = df[
                        (df.SCJ == SCJ_number) & (df.SC == SC_number) & (df["Signal"].str.contains(ASPECT))]
                    seed_results.append(len(green_signal_instances))
            all_results["Seed " + str(file_number + 1)] = pd.Series(seed_results, dtype=float)
            file_number += 1
            project_name = check_project_name(project, path)

    # Rename the indices of the results DataFrame using the Demand dependency codes
    names = [f'{SCJ}_{SC}_{ASPECT}' for SCJ, SC in zip(sites, sc)]
    all_results.index = names

    all_results["Average"] = all_results.mean(axis=1)  # Get the average per demand dependant criteria

    # Insert the results into the designated column in previously read Excel file.
    row_number = 5
    for value in all_results["Average"]:
        worksheet.cell(row=row_number, column=9).value = value
        row_number += 1

    # Append the results from all seeds into the seconds sheet
    worksheet = book["Seed validation"]
    for row in dataframe_to_rows(all_results, index=True, header=True):
        worksheet.append(row)

    # Save the file
    book.save(df_writer(project_name, "Demand_dependencies", data_directory))
